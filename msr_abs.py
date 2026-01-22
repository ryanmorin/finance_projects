import os
import re
import pandas as pd
from datetime import date
from pathlib import Path
import shutil
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, rows_from_range
from openpyxl.utils.cell import (rows_from_range, coordinate_from_string, column_index_from_string)
import warnings
import time

# -----------------------------
# Constants & config
# -----------------------------
previous_date = date(2025, 11, 30)  # USER SUPPLIED ---> THE LAST DAY OF PREVIOUS MONTH **ALWAYS**.
start_date = previous_date + pd.DateOffset(days=1)
end_date = start_date + pd.offsets.MonthEnd(0)

## ABS are ordered smallest to largest -- location of ABS name corresponds to location of Facility_ID
ABS = ["MSR_2017-1", "MSR_2019-2", "MSR_2020-2", "MSR_2023-4", "MSR_2021-2", "MSR_2018-1", "MSR_2019-1",
       "MSR_2021-3", "MSR_2020-1", "MSR_2024-1", "MSR_2017-2", "MSR_2022-3", "MSR_2023-1", "MSR_2021-1",
       "MSR_2018-2", "MSR_2022-2", "MSR_2025-1", "MSR_2023-3", "MSR_2024-2", "MSR_2022-1", "MSR_2023-2"]
FACILITY_IDS = [8, 20, 27, 56, 36, 12, 18, 39, 26, 57, 10, 50, 47, 29, 14, 43, 63, 55, 61, 40, 53]


def as_list(x):
    return list(x) if isinstance(x, (list, tuple, set)) else [x]


ALL_ABS = {w: as_list(fid) for w, fid in zip(ABS, FACILITY_IDS)}

# Regex for A1-style cell references
_CELL_REF_RE = re.compile(r"(\$?[A-Za-z]{1,3})(\$?)(\d+)")

msr_data_path = Path("/Users/ryan.morin/abs_working_folder/msr_loan_tape.csv")
txn_data_path = Path("/Users/ryan.morin/abs_working_folder/concord_transactions.csv")

## start and end columns for the two different transformations (2023-3 has more Loan Tape cols than others)
TAPE_STR_COL_IDX = column_index_from_string("A")
NOR_TAPE_END_COL_IDX = column_index_from_string("CX")
NOR_FORM_STR_COL_IDX = column_index_from_string("CY")####
NOR_FORM_END_COL_IDX = column_index_from_string("ES")
ECP_TAPE_END_COL_IDX = column_index_from_string("ES")
ECP_FORM_STR_COL_IDX = column_index_from_string("ET")###
ECP_FORM_END_COL_IDX = column_index_from_string("FX")

## Run one transformation on 'concord_transactions' tab
TXNS_STR_COL_IDX = column_index_from_string("A")
TXNS_END_COL_IDX = column_index_from_string("I")


## -----------------------------
# Load data (pandas)
## -----------------------------
msr_df_preview = pd.read_csv(msr_data_path, nrows=1)

# inconsistent dtypes out of snowflake -> force them to string for safety
dtype_map = {
    msr_df_preview.columns[10]: "string",
    msr_df_preview.columns[14]: "string",
    msr_df_preview.columns[16]: "string",
    msr_df_preview.columns[17]: "string",
    msr_df_preview.columns[43]: "string",
    msr_df_preview.columns[44]: "string",
    msr_df_preview.columns[48]: "string",
    msr_df_preview.columns[50]: "string",
    msr_df_preview.columns[70]: "string",
    msr_df_preview.columns[71]: "string",
    msr_df_preview.columns[103]: "string",
    msr_df_preview.columns[105]: "string",
    msr_df_preview.columns[130]: "string",
    msr_df_preview.columns[131]: "string",
    msr_df_preview.columns[132]: "string",
    msr_df_preview.columns[133]: "string",
    msr_df_preview.columns[135]: "string",
    msr_df_preview.columns[137]: "string",
    msr_df_preview.columns[139]: "string",
    msr_df_preview.columns[140]: "string",
    msr_df_preview.columns[141]: "string",
    msr_df_preview.columns[142]: "string",
    msr_df_preview.columns[144]: "string",
    msr_df_preview.columns[145]: "string",
}

msr_df = pd.read_csv(
    msr_data_path,
    dtype=dtype_map,
    low_memory=False,  # avoids chunked type guessing overhead
)

msr_df.columns = msr_df.columns.str.lower()

txn_df = pd.read_csv(
    txn_data_path,
    dtype=dtype_map,
    low_memory=False,  # avoids chunked type guessing overhead
)

txn_df.columns = txn_df.columns.str.lower()


# -----------------------------
# Utility functions
# -----------------------------
def format_date(dt: date) -> str:
    """Return a date formatted as MM.DD.YYYY, e.g. 10.31.2025"""
    return dt.strftime("%m.%d.%Y")


def ensure_working_folder(base_dir: str | None = "~") -> Path:
    """
    Ensure that 'abs_working_folder' exists inside base_dir.
    Returns the full Path to the folder.
    """
    base_path = Path.cwd() if base_dir is None else Path(base_dir).expanduser().resolve()
    working_folder = base_path / "abs_working_folder"
    working_folder.mkdir(exist_ok=True)
    return working_folder


def copy_change_file_name(
    file_path: Path,
    abs_name: str,
    dest_dir: Path | None = None,
) -> Path:
    """
    Copy the existing Excel file and save it with a new name based on
    abs and end_date, in dest_dir (or same directory as file_path).
    """
    if dest_dir is None:
        dest_dir = file_path.parent

    suffix = file_path.suffix

    new_file_name = f"{abs_name}_{format_date(end_date)}-WIP{suffix}"
    new_file_path = dest_dir / new_file_name

    shutil.copy2(file_path, new_file_path)

    return new_file_path


def search_for_previous_final(directory_path: Path, abs_name: str) -> Path | None:
    """
    Look for a file in directory_path whose name contains abs_name and 'final'
    and ends with '.xlsm' (case-insensitive), ignoring Excel lock files (~$).
    Return its Path or None.
    """
    candidates: list[Path] = []

    for item_name in os.listdir(directory_path):
        if item_name.startswith("~$"):  # Skip Excel lock/owner files
            continue

        lower_name = item_name.lower()
        ## its going to look for the following 3 items - if not in file name, then go to the next.
        if (
            abs_name.lower() in lower_name ## --> MAR_2017-1
            and "final" in lower_name
            and lower_name.endswith(".xlsm")
        ):
            candidates.append(directory_path / item_name)

    if not candidates:
        return None

    # Deterministic choice (alphabetical)
    return sorted(candidates)[0]


def shift_formula_rows(formula: str, row_delta: int) -> str:
    """
    Shift all *relative* row references in a formula by row_delta.
    Absolute rows (A$2, $B$3) are left unchanged.
    """

    def _repl(match: re.Match) -> str:
        col_part = match.group(1)      # e.g. "A" or "$A"
        row_dollar = match.group(2)    # "" or "$"
        row_num = int(match.group(3))  # e.g. 2

        # Absolute row -> do not change the row number
        if row_dollar == "$":
            return f"{col_part}{row_dollar}{row_num}"

        # Relative row -> shift by row_delta
        return f"{col_part}{row_num + row_delta}"

    return _CELL_REF_RE.sub(_repl, formula)


def last_nonblank_row_in_col(ws, col_idx: int, start_row: int = 2) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            last = r
    return last


def get_loan_tape_sheet(wb: openpyxl.Workbook):
    """
    Return the worksheet whose name matches 'Loan Tape' ignoring
    leading/trailing spaces and case. Raising a helpful error if not found.
    """
    target = "loan tape"
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]

    # If we reach here, nothing matched
    raise KeyError(
        f"Sheet 'Loan Tape' not found. Available sheets: {wb.sheetnames}"
    )


def get_concord_txns_sheet(wb: openpyxl.Workbook):
    """
    Return the worksheet whose name matches 'Loan Tape' ignoring
    leading/trailing spaces and case. Raising a helpful error if not found.
    """
    target = "concord_transactions"
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]

    # If we reach here, nothing matched
    raise KeyError(
        f"Sheet 'concord_transactions' not found. Available sheets: {wb.sheetnames}"
    )


def copy_range_values(src_sheet, dst_sheet, range_string: str, dest_start_cell: str = "B2") -> None:
    """
    Copy values from src_sheet[range_string] into dst_sheet starting at dest_start_cell.
    Example: range_string="B2:E25", dest_start_cell="B2".
    """
    # Parse destination start cell
    col_letters, start_row_dst = coordinate_from_string(dest_start_cell)
    start_col_dst = column_index_from_string(col_letters)

    # rows_from_range -> iterable of tuples of cell coordinates, e.g. ("B2","C2","D2","E2"), then ("B3",...)
    for row_idx, row_coords in enumerate(rows_from_range(range_string)):
        for col_idx, cell_coord in enumerate(row_coords):
            src_value = src_sheet[cell_coord].value

            target_row = start_row_dst + row_idx
            target_col = start_col_dst + col_idx

            dst_sheet.cell(row=target_row, column=target_col, value=src_value)


# -----------------------------
# Core Excel writer
# -----------------------------
def update_workbook(
    workbook_path: str | Path,
    df_ltape_for_facility: pd.DataFrame,
    df_txns_for_facility: pd.DataFrame,
) -> None:
    """
    Update workbook -> 'Loan Tape' sheet with df_ltape_for_facility rows.
    Update workbook -> 'concord_transactions' sheet with df_txns_for_facility

    - Clears existing data in A2:EY(last row).
    - Writes df_for_facility rows starting at A2.
    - Extends formulas in CZ:GX from row 2 down to last data row.
    """
    workbook_path = Path(workbook_path)

    # evals to TRUE when .xlsm
    keep_vba = workbook_path.suffix.lower() == ".xlsm"

    wb = openpyxl.load_workbook(workbook_path, keep_vba=keep_vba)

    # check the name of the ABS --> 2023-3 loan tape has more columns than others need to filter for abs name
    is_ecp = (workbook_path.name.startswith("MSR_2023-3_")) or ("MSR_2023-3" in workbook_path.name)

    # get the loan tape & transactions worksheet
    ws_ltape = get_loan_tape_sheet(wb)
    ws_txns = get_concord_txns_sheet(wb)

    # counts the number of non-null rows in the loan tape formula column
    form_start_col = ECP_FORM_STR_COL_IDX if is_ecp else NOR_FORM_STR_COL_IDX
    form_end_col = ECP_FORM_END_COL_IDX if is_ecp else NOR_FORM_END_COL_IDX
    last_formula_row = last_nonblank_row_in_col(ws_ltape, form_start_col, start_row=2)

    # beginning on line 2 and go to the last row (max_row) of both loan_tape and concord_transactions
    # line 1 is column headers
    start_row = 2
    max_row_ltape = ws_ltape.max_row
    max_row_txns = ws_txns.max_row

    if not is_ecp: ### not is_ecp means that its not 2023-3 ABS
        # --- 1. Clear existing data in loan tape A2:EY(last row) ---
        for row in ws_ltape.iter_rows(
            min_row=start_row,
            max_row=max_row_ltape,
            min_col=TAPE_STR_COL_IDX,
            max_col=NOR_TAPE_END_COL_IDX,
        ):
            for lt_cell in row:
                lt_cell.value = None

        # --- 1b. clear existing data in the concord_transactions
        for row in ws_txns.iter_rows(
            min_row=start_row,
            max_row=max_row_txns,
            min_col=TXNS_STR_COL_IDX,
            max_col=TXNS_END_COL_IDX,
        ):
            for tx_cell in row:
                tx_cell.value = None

        # --- 2. Prepare data ---
        if df_ltape_for_facility.empty:
            print("No loan tape rows found for this facility; cleared existing data only.")
            wb.save(workbook_path)
            return
        elif df_txns_for_facility.empty:
            print("No concord_transactions rows found for this facility; cleared existing data only.")
            wb.save(workbook_path)
            return

        # Convert pandas missing values in loan tape (<NA>, NaN) to Python None for Excel,
        # and use NumPy array for faster row/col access.
        df_ltape_clean = df_ltape_for_facility.astype(object).where(pd.notna(df_ltape_for_facility), None).iloc[:, :102]
        lt_values = df_ltape_clean.to_numpy()
        ln_rows, ln_cols = lt_values.shape

        df_txns_clean = df_txns_for_facility.astype(object).where(pd.notna(df_txns_for_facility), None)
        tx_values = df_txns_clean.to_numpy()
        tx_rows, tx_cols = tx_values.shape

        if ln_cols > NOR_TAPE_END_COL_IDX:
            raise ValueError(
                f"msr_df has {ln_cols} columns but only {NOR_TAPE_END_COL_IDX} columns "
                f"(A:ES) are available in the template."
            )
        elif tx_cols > TXNS_END_COL_IDX:
            raise ValueError(
                f"msr_df has {tx_cols} columns but only {TXNS_END_COL_IDX} columns "
                f"(A:I) are available in the template."
            )

        # --- 3. Write filtered data into loan tape A2:EY ---
        ws_lt_cell = ws_ltape.cell  # local alias
        base_lt_col = TAPE_STR_COL_IDX

        for i in range(ln_rows):
            row_idx = start_row + i
            lt_row_values = lt_values[i]
            for j, value in enumerate(lt_row_values):
                ws_lt_cell(row=row_idx, column=base_lt_col + j, value=value)

        last_data_row = start_row + ln_rows - 1

        # --- 4. Contract or Extend formulas in CZ:GX from row 2 to last_data_row in loan tape---
        if last_formula_row > last_data_row:
            for c in range(form_start_col, form_end_col + 1):
                for r in range(last_data_row + 1, last_formula_row + 1):
                    ws_ltape.cell(row=r, column=c).value = None
        elif last_data_row >= start_row:
            for col_idx in range(NOR_FORM_STR_COL_IDX, NOR_FORM_END_COL_IDX + 1):
                col_letter = get_column_letter(col_idx)
                origin_cell = ws_ltape[f"{col_letter}{start_row}"]
                base_formula = origin_cell.value

                # Only propagate if there is a real formula in row 2
                if not (isinstance(base_formula, str) and base_formula.startswith("=")):
                    continue

                origin_row = origin_cell.row
                for row_idx in range(start_row + 1, last_data_row + 1):
                    row_delta = row_idx - origin_row
                    formula_to_write = shift_formula_rows(base_formula, row_delta)
                    ws_ltape[f"{col_letter}{row_idx}"] = formula_to_write

        # --- 5. Write filtered data into concord_transactions A2:I ---
        ws_tx_cell = ws_txns.cell  # local alias
        base_tx_col = TXNS_STR_COL_IDX

        for i in range(tx_rows):
            row_idx = start_row + i
            tx_row_values = tx_values[i]
            for j, value in enumerate(tx_row_values):
                ws_tx_cell(row=row_idx, column=base_tx_col + j, value=value)

    else:
        # --- 1. Clear existing data in loan tape A2:EY(last row) ---
        for row in ws_ltape.iter_rows(
                min_row=start_row,
                max_row=max_row_ltape,
                min_col=TAPE_STR_COL_IDX,
                max_col=ECP_TAPE_END_COL_IDX,
        ):
            for lt_cell in row:
                lt_cell.value = None

        # --- 1b. clear existing data in the concord_transactions
        for row in ws_txns.iter_rows(
                min_row=start_row,
                max_row=max_row_txns,
                min_col=TXNS_STR_COL_IDX,
                max_col=TXNS_END_COL_IDX,
        ):
            for tx_cell in row:
                tx_cell.value = None

        # --- 2. Prepare data ---
        if df_ltape_for_facility.empty:
            print("No loan tape rows found for this facility; cleared existing data only.")
            wb.save(workbook_path)
            return
        elif df_txns_for_facility.empty:
            print("No concord_transactions rows found for this facility; cleared existing data only.")
            wb.save(workbook_path)
            return

        # Convert pandas missing values in loan tape (<NA>, NaN) to Python None for Excel,
        # and use NumPy array for faster row/col access.
        df_ltape_clean = df_ltape_for_facility.astype(object).where(pd.notna(df_ltape_for_facility), None).iloc[:, :149]
        lt_values = df_ltape_clean.to_numpy()
        ln_rows, ln_cols = lt_values.shape

        df_txns_clean = df_txns_for_facility.astype(object).where(pd.notna(df_txns_for_facility), None)
        tx_values = df_txns_clean.to_numpy()
        tx_rows, tx_cols = tx_values.shape

        if ln_cols > ECP_TAPE_END_COL_IDX:
            raise ValueError(
                f"msr_df has {ln_cols} columns but only {ECP_TAPE_END_COL_IDX} columns "
                f"(A:ES) are available in the template."
            )
        elif tx_cols > TXNS_END_COL_IDX:
            raise ValueError(
                f"msr_df has {tx_cols} columns but only {TXNS_END_COL_IDX} columns "
                f"(A:I) are available in the template."
            )

        # --- 3. Write filtered data into loan tape A2:EY ---
        ws_lt_cell = ws_ltape.cell  # local alias
        base_lt_col = TAPE_STR_COL_IDX

        for i in range(ln_rows):
            row_idx = start_row + i
            lt_row_values = lt_values[i]
            for j, value in enumerate(lt_row_values):
                ws_lt_cell(row=row_idx, column=base_lt_col + j, value=value)

        last_data_row = start_row + ln_rows - 1

        # --- 4. Extend formulas in CZ:GX from row 2 to last_data_row in loan tape---
        if last_formula_row > last_data_row:
            for c in range(form_start_col, form_end_col + 1):
                for r in range(last_data_row + 1, last_formula_row + 1):
                    ws_ltape.cell(row=r, column=c).value = None
        elif last_data_row >= start_row:
            for col_idx in range(ECP_FORM_STR_COL_IDX, ECP_FORM_END_COL_IDX + 1):
                col_letter = get_column_letter(col_idx)
                origin_cell = ws_ltape[f"{col_letter}{start_row}"]
                base_formula = origin_cell.value

                # Only propagate if there is a real formula in row 2
                if not (isinstance(base_formula, str) and base_formula.startswith("=")):
                    continue

                origin_row = origin_cell.row
                for row_idx in range(start_row + 1, last_data_row + 1):
                    row_delta = row_idx - origin_row
                    formula_to_write = shift_formula_rows(base_formula, row_delta)
                    ws_ltape[f"{col_letter}{row_idx}"] = formula_to_write

        # --- 5. Write filtered data into concord_transactions A2:I ---
        ws_tx_cell = ws_txns.cell  # local alias
        base_tx_col = TXNS_STR_COL_IDX

        for i in range(tx_rows):
            row_idx = start_row + i
            tx_row_values = tx_values[i]
            for j, value in enumerate(tx_row_values):
                ws_tx_cell(row=row_idx, column=base_tx_col + j, value=value)

    # --- Save workbook (in-place) ---
    wb.save(workbook_path)


# -----------------------------
# Driver code
# -----------------------------
if __name__ == "__main__":
    start_time = time.perf_counter()
    working_folder = ensure_working_folder()

    # Pre-split once per warehouse (handles list vs single id)
    tape_abs_to_df: dict[str, pd.DataFrame] = {}
    txn_abs_to_df: dict[str, pd.DataFrame] = {}
    for wh, fid in zip(ABS, FACILITY_IDS):
        fids = fid if isinstance(fid, (list, tuple, set)) else [fid]
        tape_abs_to_df[wh] = msr_df[msr_df["facility_id"].isin(fids)].copy()
        txn_abs_to_df[wh] = txn_df[txn_df["facility_id"].isin(fids)].copy()

    for abs_name in ABS:
        facility_ids = ALL_ABS[abs_name]  # always a list now
        df_tape_fac = tape_abs_to_df.get(abs_name, pd.DataFrame())
        df_txn_fac = txn_abs_to_df.get(abs_name, pd.DataFrame())

        prev_final_file = search_for_previous_final(
            working_folder, abs_name=abs_name
        )
        if prev_final_file is None:
            raise FileNotFoundError(
                f"No '*.xlsm' file containing '{abs_name}' and 'final' found in {working_folder}"
            )

        print(f"ABS: '{abs_name};{facility_ids}' has been started.")

        new_filepath = copy_change_file_name(
            file_path=prev_final_file,
            abs_name=abs_name,
            dest_dir=working_folder,
        )

        # update the concord_transactions and the loan tape page
        update_workbook(
            workbook_path=new_filepath,
            df_ltape_for_facility=df_tape_fac,
            df_txns_for_facility=df_txn_fac,
        )

        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        print(
            f"ABS: '{abs_name};{facility_ids}' has been completed. "
            f"Time to complete: {elapsed_time / 60:.4f} minutes"
        )
