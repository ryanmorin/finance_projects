import os
import re
import pandas as pd
from datetime import date
from pathlib import Path
import shutil
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, rows_from_range
from openpyxl.utils.cell import (rows_from_range, coordinate_from_string, column_index_from_string)
import warnings
import time

# -----------------------------
# Constants & config
# -----------------------------
previous_date = date(2025, 12, 31)  # USER SUPPLIED ---> THE LAST DAY OF PREVIOUS MONTH **ALWAYS**.
start_date = previous_date + pd.DateOffset(days=1)
end_date = start_date + pd.offsets.MonthEnd(0)

WAREHOUSES = ["SOLAR_MOSAIC_LLC", "MF7", "MF5", "MF6", "MF9", "MF8", "MF1"]        # USER SUPPLIED
FACILITY_IDS = [[1, 48], 45, 42, 44, 51, 49, 4]                                    # USER SUPPLIED


# Map warehouse -> facility_id
def as_list(x):
    return list(x) if isinstance(x, (list, tuple, set)) else [x]


ALL_FACILITIES = {w: as_list(fid) for w, fid in zip(WAREHOUSES, FACILITY_IDS)}

# Regex for A1-style cell references
_CELL_REF_RE = re.compile(r"(\$?[A-Za-z]{1,3})(\$?)(\d+)")

msr_data_path = Path("/Users/ryan.morin/msr_working_folder/msr_loan_tape.csv")

# Precompute Excel column indices so we don't do it in hot loops
DATA_START_COL_IDX = column_index_from_string("A") ## data start for loan tape column
DATA_END_COL_IDX = column_index_from_string("EY") ## data end for loan tape column
FORMULA_START_COL_IDX = column_index_from_string("CZ")
FORMULA_END_COL_IDX = column_index_from_string("GY")  # MF9 GM & MF1 GY
IO_SOURCE_RANGE = 'B2:F28' ## -- the full range of the input and output data range

# -----------------------------
# Load data (pandas)
# -----------------------------
msr_df_preview = pd.read_csv(msr_data_path, nrows=1)

# inconsistent dtypes out of snowflake -> force them to string for safety
dtype_map = {
    msr_df_preview.columns[10]: "string",
    msr_df_preview.columns[14]: "string",
    msr_df_preview.columns[16]: "string",
    msr_df_preview.columns[17]: "string",
    msr_df_preview.columns[43]: "string",
    msr_df_preview.columns[44]: "string",
    msr_df_preview.columns[48]: "string",
    msr_df_preview.columns[50]: "string",
    msr_df_preview.columns[70]: "string",
    msr_df_preview.columns[71]: "string",
    msr_df_preview.columns[103]: "string",
    msr_df_preview.columns[105]: "string",
    msr_df_preview.columns[130]: "string",
    msr_df_preview.columns[131]: "string",
    msr_df_preview.columns[132]: "string",
    msr_df_preview.columns[133]: "string",
    msr_df_preview.columns[135]: "string",
    msr_df_preview.columns[137]: "string",
    msr_df_preview.columns[139]: "string",
    msr_df_preview.columns[140]: "string",
    msr_df_preview.columns[141]: "string",
    msr_df_preview.columns[142]: "string",
    msr_df_preview.columns[144]: "string",
    msr_df_preview.columns[145]: "string",
}

msr_df = pd.read_csv(
    msr_data_path,
    dtype=dtype_map,
    low_memory=False,  # avoids chunked type guessing overhead
)

msr_df.columns = msr_df.columns.str.lower()

# overcome the print area warning from excel
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=r"Print area cannot be set to Defined name: Table1\[#All\].*",
)


# -----------------------------
# Utility functions
# -----------------------------
def format_date(dt: date) -> str:
    """Return a date formatted as MM.DD.YYYY, e.g. 10.31.2025"""
    return dt.strftime("%m.%d.%Y")


def ensure_working_folder(base_dir: str | None = "~") -> Path:
    """
    Ensure that 'msr_working_folder' exists inside base_dir.
    Returns the full Path to the folder.
    """
    base_path = Path.cwd() if base_dir is None else Path(base_dir).expanduser().resolve()
    working_folder = base_path / "msr_working_folder"
    working_folder.mkdir(exist_ok=True)
    return working_folder


def copy_change_file_name(
    file_path: Path,
    warehouse_name: str,
    dest_dir: Path | None = None,
) -> Path:
    """
    Copy the existing Excel file and save it with a new name based on
    warehouse and end_date, in dest_dir (or same directory as file_path).
    """
    if dest_dir is None:
        dest_dir = file_path.parent

    suffix = file_path.suffix

    new_file_name = f"{warehouse_name}_MSR {format_date(end_date)} - WIP{suffix}"
    new_file_path = dest_dir / new_file_name

    shutil.copy2(file_path, new_file_path)

    #print(f'Excel file "{file_path}" successfully copied and renamed to "{new_file_path}"')
    return new_file_path


def search_for_previous_final(directory_path: Path, warehouse_name: str) -> Path | None:
    """
    Look for a file in directory_path whose name contains warehouse_name and 'final'
    and ends with '.xlsm' (case-insensitive), ignoring Excel lock files (~$).
    Return its Path or None.
    """
    candidates: list[Path] = []

    for item_name in os.listdir(directory_path):
        if item_name.startswith("~$"):  # Skip Excel lock/owner files
            continue

        lower_name = item_name.lower()
        ## its going to look for the following 3 items - if not in file name, then go to the next.
        if (
            warehouse_name.lower() in lower_name ## --> eg MF9
            and "final" in lower_name
            and lower_name.endswith(".xlsm")
        ):
            candidates.append(directory_path / item_name)

    if not candidates:
        return None

    # Deterministic choice (alphabetical); you could switch to "latest by mtime" if desired
    return sorted(candidates)[0]


def shift_formula_rows(formula: str, row_delta: int) -> str:
    """
    Shift all *relative* row references in a formula by row_delta.
    Absolute rows (A$2, $B$3) are left unchanged.
    """

    def _repl(match: re.Match) -> str:
        col_part = match.group(1)      # e.g. "A" or "$A"
        row_dollar = match.group(2)    # "" or "$"
        row_num = int(match.group(3))  # e.g. 2

        # Absolute row -> do not change the row number
        if row_dollar == "$":
            return f"{col_part}{row_dollar}{row_num}"

        # Relative row -> shift by row_delta
        return f"{col_part}{row_num + row_delta}"

    return _CELL_REF_RE.sub(_repl, formula)


def get_loan_tape_sheet(wb: openpyxl.Workbook):
    """
    Return the worksheet whose name matches 'Loan Tape' ignoring
    leading/trailing spaces and case. Raise a helpful error if not found.
    """
    target = "loan tape"
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]

    # If we reach here, nothing matched
    raise KeyError(
        f"Sheet 'Loan Tape' not found. Available sheets: {wb.sheetnames}"
    )


def get_output_sheet(wb: openpyxl.Workbook):
    """
    Return the worksheet whose name matches 'output' ignoring
    leading/trailing spaces and case. Raise a helpful error if not found.
    """
    target = "output"
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]

    # If we reach here, nothing matched
    raise KeyError(
        f"Sheet 'output' not found. Available sheets: {wb.sheetnames}"
    )


def get_input_sheet(wb: openpyxl.Workbook):
    """
    Return the worksheet whose name matches 'output' ignoring
    leading/trailing spaces and case. Raise a helpful error if not found.
    """
    target = "input"
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]

    # If we reach here, nothing matched
    raise KeyError(
        f"Sheet 'input' not found. Available sheets: {wb.sheetnames}"
    )


def copy_range_values(src_sheet, dst_sheet, range_string: str, dest_start_cell: str = "B2") -> None:
    """
    Copy values from src_sheet[range_string] into dst_sheet starting at dest_start_cell.
    Example: range_string="B2:E25", dest_start_cell="B2".
    """
    # Parse destination start cell
    col_letters, start_row_dst = coordinate_from_string(dest_start_cell)
    start_col_dst = column_index_from_string(col_letters)

    # rows_from_range -> iterable of tuples of cell coordinates, e.g. ("B2","C2","D2","E2"), then ("B3",...)
    for row_idx, row_coords in enumerate(rows_from_range(range_string)):
        for col_idx, cell_coord in enumerate(row_coords):
            src_value = src_sheet[cell_coord].value

            target_row = start_row_dst + row_idx
            target_col = start_col_dst + col_idx

            dst_sheet.cell(row=target_row, column=target_col, value=src_value)


# -----------------------------
# Core Excel writer
# -----------------------------
def update_workbook(
    workbook_path: str | Path,
    df_for_facility: pd.DataFrame,
) -> None:
    """
    Update workbook -> 'Loan Tape' sheet with df_for_facility rows.

    - Clears existing data in A2:EY(last row).
    - Writes df_for_facility rows starting at A2.
    - Extends formulas in CZ:GX from row 2 down to last data row.
    """
    workbook_path = Path(workbook_path)

    # evals to TRUE when .xlsm
    keep_vba = workbook_path.suffix.lower() == '.xlsm'

    wb = openpyxl.load_workbook(workbook_path, keep_vba=keep_vba)

    # get the loan tape worksheet
    ws_ltape = get_loan_tape_sheet(wb)

    # beginning on line 2 and go to the last row (max_row) - line 1 is column headers
    start_row = 2
    max_row_ltape = ws_ltape.max_row

    # --- 1. Clear existing data in loan tape A2:EY(last row) ---
    for row in ws_ltape.iter_rows(
        min_row=start_row,
        max_row=max_row_ltape,
        min_col=DATA_START_COL_IDX,
        max_col=DATA_END_COL_IDX,
    ):
        for cell in row:
            cell.value = None

    # --- 2. Prepare data (df_for_facility is already filtered) ---
    if df_for_facility.empty:
        print("No rows found for this facility; cleared existing data only.")
        wb.save(workbook_path)
        return

    # Convert pandas missing values in loan tape (<NA>, NaN) to Python None for Excel,
    # and use NumPy array for faster row/col access.
    df_clean = df_for_facility.astype(object).where(pd.notna(df_for_facility), None)
    values = df_clean.to_numpy()
    n_rows, n_cols = values.shape

    if n_cols > DATA_END_COL_IDX:
        raise ValueError(
            f"msr_df has {n_cols} columns but only {DATA_END_COL_IDX} columns "
            f"(A:EY) are available in the template."
        )

    # --- 3. Write filtered data into loan tape A2:EY ---
    ws_cell = ws_ltape.cell  # local alias
    base_col = DATA_START_COL_IDX

    for i in range(n_rows):
        row_idx = start_row + i
        row_values = values[i]
        for j, value in enumerate(row_values):
            ws_cell(row=row_idx, column=base_col + j, value=value)

    last_data_row = start_row + n_rows - 1

    # --- 4. Extend formulas in CZ:GX from row 2 to last_data_row in loan tape---
    if last_data_row >= start_row:
        for col_idx in range(FORMULA_START_COL_IDX, FORMULA_END_COL_IDX + 1):
            col_letter = get_column_letter(col_idx)
            origin_cell = ws_ltape[f"{col_letter}{start_row}"]
            base_formula = origin_cell.value

            # Only propagate if there is a real formula in row 2
            if not (isinstance(base_formula, str) and base_formula.startswith("=")):
                continue

            origin_row = origin_cell.row
            for row_idx in range(start_row + 1, last_data_row + 1):
                row_delta = row_idx - origin_row
                formula_to_write = shift_formula_rows(base_formula, row_delta)
                ws_ltape[f"{col_letter}{row_idx}"] = formula_to_write

    # --- Save workbook (in-place) ---
    wb.save(workbook_path)


# -----------------------------
# Driver code
# -----------------------------
if __name__ == "__main__":
    start_time = time.perf_counter()
    working_folder = ensure_working_folder()

    # Pre-split once per warehouse (handles list vs single id)
    warehouse_to_df: dict[str, pd.DataFrame] = {}
    for wh, fid in zip(WAREHOUSES, FACILITY_IDS):
        fids = fid if isinstance(fid, (list, tuple, set)) else [fid]
        warehouse_to_df[wh] = msr_df[msr_df["facility_id"].isin(fids)].copy()

    for warehouse_name in WAREHOUSES:
        facility_ids = ALL_FACILITIES[warehouse_name]  # always a list now
        df_fac = warehouse_to_df.get(warehouse_name, pd.DataFrame())

        prev_final_file = search_for_previous_final(
            working_folder, warehouse_name=warehouse_name
        )
        if prev_final_file is None:
            raise FileNotFoundError(
                f"No '*.xlsm' file containing '{warehouse_name}' and 'final' found in {working_folder}"
            )

        print(f"Warehouse: '{warehouse_name};{facility_ids}' has been started.")

        new_filepath = copy_change_file_name(
            file_path=prev_final_file,
            warehouse_name=warehouse_name,
            dest_dir=working_folder,
        )

        source_wb = openpyxl.load_workbook(filename=prev_final_file, data_only=True, read_only=True)
        source_ws = get_output_sheet(source_wb)

        keep_vba_dest = new_filepath.suffix.lower() == ".xlsm"
        dest_wb = openpyxl.load_workbook(filename=new_filepath, keep_vba=keep_vba_dest)
        dest_ws = get_input_sheet(dest_wb)

        copy_range_values(src_sheet=source_ws, dst_sheet=dest_ws, range_string=IO_SOURCE_RANGE)
        dest_wb.save(new_filepath)

        # Update loan tape using the filtered df for this warehouse’s facility id(s)
        update_workbook(workbook_path=new_filepath, df_for_facility=df_fac)

        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        print(
            f"Warehouse: '{warehouse_name};{facility_ids}' has been completed. "
            f"Time to complete: {elapsed_time / 60:.4f} minutes"
        )